"""
failure_analyzer.py — Automated Test Failure Analysis Tool
============================================================
Parses log files to extract, categorize, and report test failures.
Outputs:  Excel workbook (multi-sheet) + interactive HTML report.

Usage:
    python failure_analyzer.py --path logs/
    python failure_analyzer.py --path run.log
    python failure_analyzer.py --path logs/ --pattern "ERROR|FAIL|CRITICAL"
    python failure_analyzer.py --path logs/ --output my_report
"""

import os
import re
import sys
import json
import argparse
import logging
from datetime import datetime
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# ─────────────────────────── LOGGING ────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ─────────────────────────── CONSTANTS ──────────────────────────
DEFAULT_OUTPUT_STEM = "failure_report"
SUPPORTED_EXTENSIONS = {".log", ".txt", ".out"}

FAILURE_CATEGORIES = {
    "Timeout":      re.compile(r"timeout|timed?\s*out", re.I),
    "Assertion":    re.compile(r"assert(ion)?(\s+error)?|expected\s+.+\s+but\s+got", re.I),
    "Connection":   re.compile(r"connect(ion)?\s*(refused|error|failed)|socket\s+error", re.I),
    "NullRef":      re.compile(r"null\s*(pointer|reference|deref)|none\s*type", re.I),
    "Permission":   re.compile(r"permission\s+denied|access\s+denied|unauthori[sz]ed", re.I),
    "Memory":       re.compile(r"out\s+of\s+memory|memory\s+(error|leak|overflow)|oom", re.I),
    "Syntax":       re.compile(r"syntax\s+error|invalid\s+syntax|parse\s+error", re.I),
    "FileNotFound": re.compile(r"file\s+not\s+found|no\s+such\s+file|path\s+does\s+not\s+exist", re.I),
}

# Excel palette
C = {
    "header_bg":  "1F3864",   # dark navy
    "header_fg":  "FFFFFF",
    "fail_bg":    "FFD7D7",
    "fail_fg":    "C00000",
    "warn_bg":    "FFF2CC",
    "pass_bg":    "E2EFDA",
    "alt_row":    "F5F8FF",
    "border":     "BDD7EE",
    "summary_bg": "D6E4F0",
    "accent":     "2E75B6",
}


# ─────────────────────────── PARSING ────────────────────────────
# Default pattern — matches lines like:
#   [2024-01-15 10:32:01] Test Failed: Test: login_flow - reason: timeout
#   FAIL  test_api_response  (0.003s)
#   ERROR  TestDatabase.test_connect  AssertionError
DEFAULT_FAILURE_RE = re.compile(
    r"(?P<timestamp>\d{4}-\d{2}-\d{2}[\sT]\d{2}:\d{2}:\d{2})?"
    r".*?"
    r"(?:FAIL(?:ED|URE)?|ERROR|Test\s+Failed)[:\s]+"
    r"(?:Test:\s*)?(?P<test_name>[\w\.:_\-\/]+)"
    r"(?:\s*[-–]\s*(?:reason:\s*)?(?P<reason>.+?))?$",
    re.I,
)


def categorize(text: str) -> str:
    for cat, pattern in FAILURE_CATEGORIES.items():
        if pattern.search(text):
            return cat
    return "Other"


def extract_timestamp(raw: str) -> str | None:
    m = re.search(r"\d{4}-\d{2}-\d{2}[\sT]\d{2}:\d{2}:\d{2}", raw)
    return m.group(0) if m else None


def parse_log(log_path: str, custom_re: re.Pattern | None = None) -> list[dict]:
    """
    Parse a single log file and return a list of failure records.
    Each record: {test_name, reason, category, raw_line, line_no, log_file, log_ts}
    """
    path = Path(log_path)
    if not path.exists():
        log.error("Log not found: %s", log_path)
        return []

    pattern = custom_re or DEFAULT_FAILURE_RE
    seen: set[str] = set()
    records: list[dict] = []

    with open(path, "r", errors="replace") as fh:
        for lineno, line in enumerate(fh, 1):
            line = line.rstrip()
            m = pattern.search(line)
            if not m:
                continue

            groups = m.groupdict()
            test_name = (groups.get("test_name") or "Unknown_Test").strip()
            reason    = (groups.get("reason")    or "").strip()
            log_ts    = groups.get("timestamp") or extract_timestamp(line) or ""

            # Deduplicate within same file
            key = f"{test_name}::{reason}"
            if key in seen:
                continue
            seen.add(key)

            records.append({
                "test_name": test_name,
                "reason":    reason or "—",
                "category":  categorize(f"{test_name} {reason} {line}"),
                "raw_line":  line[:300],
                "line_no":   lineno,
                "log_file":  path.name,
                "log_path":  str(path),
                "log_ts":    log_ts,
            })

    return records


# ─────────────────────────── EXCEL ──────────────────────────────
def _thin_border():
    s = Side(style="thin", color=C["border"])
    return Border(left=s, right=s, top=s, bottom=s)


def _header_cell(ws, row, col, value, width=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=True, color=C["header_fg"], size=11)
    cell.fill      = PatternFill("solid", fgColor=C["header_bg"])
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _thin_border()
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return cell


def _data_cell(ws, row, col, value, alt=False, red=False, bold=False):
    cell = ws.cell(row=row, column=col, value=value)
    if red:
        cell.fill = PatternFill("solid", fgColor=C["fail_bg"])
        cell.font = Font(color=C["fail_fg"], bold=bold)
    elif alt:
        cell.fill = PatternFill("solid", fgColor=C["alt_row"])
        cell.font = Font(bold=bold)
    else:
        cell.font = Font(bold=bold)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border    = _thin_border()
    return cell


def build_excel(all_records: list[dict], output_path: str):
    wb = Workbook()

    # ── Sheet 1: All Failures ────────────────────────────────────
    ws1 = wb.active
    ws1.title = "All_Failures"
    ws1.row_dimensions[1].height = 30

    headers1 = ["#", "Test Name", "Category", "Reason / Message",
                 "Log File", "Line No.", "Log Timestamp", "Parsed At"]
    widths1  = [5, 30, 15, 45, 25, 8, 20, 20]
    for c, (h, w) in enumerate(zip(headers1, widths1), 1):
        _header_cell(ws1, 1, c, h, w)

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for i, r in enumerate(all_records, 1):
        alt = (i % 2 == 0)
        _data_cell(ws1, i+1, 1, i,              alt=alt, red=True)
        _data_cell(ws1, i+1, 2, r["test_name"], alt=alt, red=True, bold=True)
        _data_cell(ws1, i+1, 3, r["category"],  alt=alt, red=True)
        _data_cell(ws1, i+1, 4, r["reason"],    alt=alt, red=True)
        _data_cell(ws1, i+1, 5, r["log_file"],  alt=alt, red=True)
        _data_cell(ws1, i+1, 6, r["line_no"],   alt=alt, red=True)
        _data_cell(ws1, i+1, 7, r["log_ts"],    alt=alt, red=True)
        _data_cell(ws1, i+1, 8, now_str,        alt=alt, red=True)

    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = ws1.dimensions

    # ── Sheet 2: By Category ─────────────────────────────────────
    ws2 = wb.create_sheet("By_Category")
    ws2.row_dimensions[1].height = 30

    cat_map: dict[str, list[dict]] = defaultdict(list)
    for r in all_records:
        cat_map[r["category"]].append(r)

    headers2 = ["Category", "Count", "% of Total", "Sample Tests"]
    widths2  = [18, 10, 14, 55]
    for c, (h, w) in enumerate(zip(headers2, widths2), 1):
        _header_cell(ws2, 1, c, h, w)

    total = len(all_records) or 1
    for i, (cat, recs) in enumerate(sorted(cat_map.items(), key=lambda x: -len(x[1])), 2):
        pct = f"{len(recs)/total*100:.1f}%"
        samples = ", ".join(dict.fromkeys(r["test_name"] for r in recs))[:120]
        alt = (i % 2 == 0)
        _data_cell(ws2, i, 1, cat,       alt=alt, bold=True)
        _data_cell(ws2, i, 2, len(recs), alt=alt)
        _data_cell(ws2, i, 3, pct,       alt=alt)
        _data_cell(ws2, i, 4, samples,   alt=alt)

    # Bar chart
    chart = BarChart()
    chart.type  = "col"
    chart.title = "Failures by Category"
    chart.y_axis.title = "Count"
    data_ref   = Reference(ws2, min_col=2, min_row=1, max_row=len(cat_map)+1)
    cats_ref   = Reference(ws2, min_col=1, min_row=2, max_row=len(cat_map)+1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4
    chart.width  = 22
    chart.height = 14
    ws2.add_chart(chart, f"F2")

    # ── Sheet 3: Per-Log Summary ──────────────────────────────────
    ws3 = wb.create_sheet("Per_Log_Summary")
    ws3.row_dimensions[1].height = 30

    log_map: dict[str, list[dict]] = defaultdict(list)
    for r in all_records:
        log_map[r["log_path"]].append(r)

    headers3 = ["Log File", "Full Path", "Total Failures",
                 "Unique Tests", "Categories Hit", "Severity"]
    widths3  = [25, 50, 14, 14, 25, 12]
    for c, (h, w) in enumerate(zip(headers3, widths3), 1):
        _header_cell(ws3, 1, c, h, w)

    for i, (lp, recs) in enumerate(log_map.items(), 2):
        cats   = ", ".join(sorted(set(r["category"] for r in recs)))
        unique = len(set(r["test_name"] for r in recs))
        sev    = "HIGH" if len(recs) > 10 else ("MEDIUM" if len(recs) > 3 else "LOW")
        alt    = (i % 2 == 0)
        _data_cell(ws3, i, 1, Path(lp).name, alt=alt, bold=True)
        _data_cell(ws3, i, 2, lp,            alt=alt)
        _data_cell(ws3, i, 3, len(recs),     alt=alt)
        _data_cell(ws3, i, 4, unique,        alt=alt)
        _data_cell(ws3, i, 5, cats,          alt=alt)
        cell = _data_cell(ws3, i, 6, sev, alt=alt, bold=True)
        if sev == "HIGH":
            cell.fill = PatternFill("solid", fgColor="FFD7D7")
            cell.font = Font(color="C00000", bold=True)
        elif sev == "MEDIUM":
            cell.fill = PatternFill("solid", fgColor="FFF2CC")
            cell.font = Font(color="7F6000", bold=True)
        else:
            cell.fill = PatternFill("solid", fgColor="E2EFDA")
            cell.font = Font(color="375623", bold=True)

    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = ws3.dimensions

    # ── Sheet 4: Executive Summary ────────────────────────────────
    ws4 = wb.create_sheet("Summary")
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 30
    ws4.column_dimensions["B"].width = 25

    def kv(r, label, value, bold_val=False):
        lc = ws4.cell(row=r, column=1, value=label)
        vc = ws4.cell(row=r, column=2, value=value)
        lc.font = Font(bold=True, size=11, color=C["accent"])
        vc.font = Font(bold=bold_val, size=11)
        lc.fill = vc.fill = PatternFill("solid", fgColor=C["summary_bg"])
        lc.border = vc.border = _thin_border()
        lc.alignment = Alignment(vertical="center")
        vc.alignment = Alignment(vertical="center")
        ws4.row_dimensions[r].height = 22

    title_cell = ws4.cell(row=1, column=1, value="🔍  Failure Analysis — Executive Summary")
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor=C["header_bg"])
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws4.merge_cells("A1:B1")
    ws4.row_dimensions[1].height = 36

    kv(3,  "Report Generated",  datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    kv(4,  "Total Log Files",   len(log_map))
    kv(5,  "Total Failures",    len(all_records), bold_val=True)
    kv(6,  "Unique Tests",      len(set(r["test_name"] for r in all_records)))
    kv(7,  "Failure Categories",", ".join(sorted(cat_map.keys())))
    kv(8,  "Top Category",      max(cat_map, key=lambda c: len(cat_map[c])) if cat_map else "—")
    kv(9,  "Highest-Risk Log",  max(log_map, key=lambda l: len(log_map[l]), default="—"))

    wb.save(output_path)
    log.info("Excel report saved → %s", output_path)


# ─────────────────────────── HTML ───────────────────────────────
def build_html(all_records: list[dict], output_path: str, source_path: str):
    cat_map: dict[str, int] = defaultdict(int)
    for r in all_records:
        cat_map[r["category"]] += 1

    log_map: dict[str, list[dict]] = defaultdict(list)
    for r in all_records:
        log_map[r["log_file"]].append(r)

    cat_json   = json.dumps(dict(cat_map))
    log_labels = json.dumps(list(log_map.keys()))
    log_counts = json.dumps([len(v) for v in log_map.values()])

    rows = ""
    for i, r in enumerate(all_records):
        badge_color = {
            "Timeout":      "#e67e22",
            "Assertion":    "#e74c3c",
            "Connection":   "#9b59b6",
            "NullRef":      "#c0392b",
            "Permission":   "#d35400",
            "Memory":       "#8e44ad",
            "Syntax":       "#2980b9",
            "FileNotFound": "#16a085",
        }.get(r["category"], "#7f8c8d")

        rows += f"""
        <tr>
          <td class="mono">{i+1}</td>
          <td><strong>{r['test_name']}</strong></td>
          <td><span class="badge" style="background:{badge_color}">{r['category']}</span></td>
          <td class="mono small">{r['reason']}</td>
          <td class="mono small">{r['log_file']}</td>
          <td class="mono small center">{r['line_no']}</td>
          <td class="mono small">{r['log_ts']}</td>
        </tr>"""

    summary_cards = ""
    metrics = [
        ("📁 Logs Scanned",     len(log_map),                                  "#3498db"),
        ("❌ Total Failures",   len(all_records),                               "#e74c3c"),
        ("🔬 Unique Tests",     len(set(r["test_name"] for r in all_records)), "#9b59b6"),
        ("🗂️ Categories",       len(cat_map),                                   "#e67e22"),
    ]
    for label, val, color in metrics:
        summary_cards += f"""
        <div class="card" style="border-top:4px solid {color}">
          <div class="card-val" style="color:{color}">{val}</div>
          <div class="card-lbl">{label}</div>
        </div>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1.0"/>
  <title>Failure Analysis Report</title>
  <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
  <style>
    :root{{--navy:#1F3864;--red:#e74c3c;--bg:#f0f4f8;--card:#fff;--text:#2c3e50;--mono:#2d3436}}
    *{{box-sizing:border-box;margin:0;padding:0}}
    body{{font-family:'Segoe UI',sans-serif;background:var(--bg);color:var(--text);font-size:14px}}
    header{{background:var(--navy);color:#fff;padding:24px 40px;display:flex;align-items:center;justify-content:space-between}}
    header h1{{font-size:22px;font-weight:700;letter-spacing:.5px}}
    header .meta{{font-size:12px;opacity:.75;text-align:right}}
    .container{{max-width:1400px;margin:0 auto;padding:24px 32px}}
    .cards{{display:flex;gap:16px;flex-wrap:wrap;margin-bottom:28px}}
    .card{{background:var(--card);border-radius:10px;padding:20px 28px;flex:1;min-width:160px;
           box-shadow:0 2px 8px rgba(0,0,0,.08)}}
    .card-val{{font-size:36px;font-weight:800;line-height:1}}
    .card-lbl{{font-size:12px;color:#7f8c8d;margin-top:6px;font-weight:600;text-transform:uppercase;letter-spacing:.5px}}
    .charts-row{{display:flex;gap:20px;flex-wrap:wrap;margin-bottom:28px}}
    .chart-box{{background:var(--card);border-radius:10px;padding:20px;flex:1;min-width:300px;
                box-shadow:0 2px 8px rgba(0,0,0,.08)}}
    .chart-box h3{{font-size:14px;font-weight:700;color:var(--navy);margin-bottom:14px;text-transform:uppercase;letter-spacing:.4px}}
    .table-wrap{{background:var(--card);border-radius:10px;box-shadow:0 2px 8px rgba(0,0,0,.08);overflow:hidden}}
    .table-header{{display:flex;align-items:center;justify-content:space-between;padding:16px 20px;border-bottom:1px solid #ecf0f1}}
    .table-header h3{{font-size:14px;font-weight:700;color:var(--navy);text-transform:uppercase;letter-spacing:.4px}}
    input#search{{border:1px solid #dfe6e9;border-radius:6px;padding:7px 14px;font-size:13px;width:240px;outline:none}}
    input#search:focus{{border-color:#3498db}}
    table{{width:100%;border-collapse:collapse}}
    thead{{background:var(--navy)}}
    thead th{{color:#fff;padding:11px 14px;text-align:left;font-size:12px;font-weight:700;letter-spacing:.4px;white-space:nowrap}}
    tbody tr{{border-bottom:1px solid #f0f0f0;transition:background .15s}}
    tbody tr:hover{{background:#eaf4fb}}
    tbody td{{padding:10px 14px;vertical-align:middle}}
    .mono{{font-family:'Consolas','Courier New',monospace}}
    .small{{font-size:12px;color:#636e72}}
    .center{{text-align:center}}
    .badge{{display:inline-block;padding:3px 10px;border-radius:20px;font-size:11px;font-weight:700;color:#fff;white-space:nowrap}}
    footer{{text-align:center;padding:20px;font-size:12px;color:#b2bec3}}
    @media(max-width:700px){{.charts-row{{flex-direction:column}}.cards{{flex-direction:column}}}}
  </style>
</head>
<body>
<header>
  <div>
    <h1>🔍 Failure Analysis Report</h1>
    <div style="font-size:13px;opacity:.8;margin-top:4px">Source: <code>{source_path}</code></div>
  </div>
  <div class="meta">
    Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}<br>
    {len(all_records)} failures · {len(log_map)} log files
  </div>
</header>

<div class="container">
  <!-- KPI Cards -->
  <div class="cards">{summary_cards}</div>

  <!-- Charts -->
  <div class="charts-row">
    <div class="chart-box" style="max-width:420px">
      <h3>Failures by Category</h3>
      <canvas id="catChart" height="200"></canvas>
    </div>
    <div class="chart-box">
      <h3>Failures per Log File</h3>
      <canvas id="logChart" height="200"></canvas>
    </div>
  </div>

  <!-- Table -->
  <div class="table-wrap">
    <div class="table-header">
      <h3>All Failures ({len(all_records)})</h3>
      <input id="search" type="text" placeholder="🔍  Filter tests…" oninput="filterTable(this.value)"/>
    </div>
    <table id="failTable">
      <thead>
        <tr>
          <th>#</th><th>Test Name</th><th>Category</th><th>Reason / Message</th>
          <th>Log File</th><th>Line</th><th>Timestamp</th>
        </tr>
      </thead>
      <tbody>{rows}</tbody>
    </table>
  </div>
</div>

<footer>failure_analyzer.py — Auto-generated on {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</footer>

<script>
// Category doughnut
const catData = {cat_json};
new Chart(document.getElementById('catChart'), {{
  type:'doughnut',
  data:{{
    labels: Object.keys(catData),
    datasets:[{{
      data: Object.values(catData),
      backgroundColor:['#e74c3c','#e67e22','#f1c40f','#2ecc71','#3498db','#9b59b6','#1abc9c','#7f8c8d'],
      borderWidth:2,borderColor:'#fff'
    }}]
  }},
  options:{{plugins:{{legend:{{position:'right',labels:{{font:{{size:12}}}}}}}}}}
}});

// Log-file bar chart
new Chart(document.getElementById('logChart'), {{
  type:'bar',
  data:{{
    labels: {log_labels},
    datasets:[{{
      label:'Failures',
      data: {log_counts},
      backgroundColor:'rgba(52,152,219,.75)',
      borderColor:'#2980b9',
      borderWidth:1,
      borderRadius:4
    }}]
  }},
  options:{{
    responsive:true,
    plugins:{{legend:{{display:false}}}},
    scales:{{y:{{beginAtZero:true,ticks:{{stepSize:1}}}},x:{{ticks:{{maxRotation:40,font:{{size:11}}}}}}}}
  }}
}});

// Live search
function filterTable(q) {{
  const rows = document.querySelectorAll('#failTable tbody tr');
  q = q.toLowerCase();
  rows.forEach(r => {{
    r.style.display = r.textContent.toLowerCase().includes(q) ? '' : 'none';
  }});
}}
</script>
</body>
</html>"""

    with open(output_path, "w", encoding="utf-8") as fh:
        fh.write(html)
    log.info("HTML  report saved → %s", output_path)


# ─────────────────────────── MAIN ───────────────────────────────
def collect_logs(path: str) -> list[str]:
    p = Path(path)
    if p.is_file():
        return [str(p)]
    if p.is_dir():
        return [
            str(f) for f in sorted(p.rglob("*"))
            if f.is_file() and f.suffix.lower() in SUPPORTED_EXTENSIONS
        ]
    log.error("Path not found: %s", path)
    return []


def main():
    parser = argparse.ArgumentParser(
        description="Failure Analyzer — parse logs, report failures",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--path",    required=True,              help="Log file or folder")
    parser.add_argument("--output",  default=DEFAULT_OUTPUT_STEM, help="Output filename stem (no extension)")
    parser.add_argument("--pattern", default=None,               help="Custom regex failure pattern")
    parser.add_argument("--no-html", action="store_true",        help="Skip HTML output")
    parser.add_argument("--no-excel",action="store_true",        help="Skip Excel output")
    args = parser.parse_args()

    custom_re = re.compile(args.pattern, re.I) if args.pattern else None

    log_files = collect_logs(args.path)
    if not log_files:
        log.error("No log files found at: %s", args.path)
        sys.exit(1)

    log.info("Found %d log file(s) to process.", len(log_files))

    all_records: list[dict] = []
    for lp in log_files:
        records = parse_log(lp, custom_re)
        log.info("  %-45s  →  %d failure(s)", Path(lp).name, len(records))
        all_records.extend(records)

    if not all_records:
        log.warning("No failures detected across all logs.")
        sys.exit(0)

    xlsx_path = f"{args.output}.xlsx"
    html_path = f"{args.output}.html"

    if not args.no_excel:
        build_excel(all_records, xlsx_path)
    if not args.no_html:
        build_html(all_records, html_path, args.path)

    # ── Console summary ──────────────────────────────────────────
    cat_map: dict[str, int] = defaultdict(int)
    for r in all_records:
        cat_map[r["category"]] += 1

    print("\n" + "═"*52)
    print("  FAILURE ANALYSIS SUMMARY")
    print("═"*52)
    print(f"  Logs processed   : {len(log_files)}")
    print(f"  Total failures   : {len(all_records)}")
    print(f"  Unique tests     : {len(set(r['test_name'] for r in all_records))}")
    print("  ─────────────────────────────────────────────")
    for cat, cnt in sorted(cat_map.items(), key=lambda x: -x[1]):
        bar = "█" * min(cnt, 30)
        print(f"  {cat:<14} {cnt:>4}  {bar}")
    print("═"*52)
    if not args.no_excel:
        print(f"  Excel  → {xlsx_path}")
    if not args.no_html:
        print(f"  HTML   → {html_path}")
    print("═"*52 + "\n")


if __name__ == "__main__":
    main()
